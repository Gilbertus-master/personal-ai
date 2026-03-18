import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


MAX_ROWS_PER_SHEET = 200
MAX_COLS_PER_SHEET = 30


@dataclass
class ParsedSpreadsheet:
    title: str
    created_at: Optional[datetime]
    author: Optional[str]
    participants: list[str]
    raw_path: str
    text: str
    file_type: str


def _clean_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _sheet_to_text(ws) -> str:
    lines = [f"=== SHEET: {ws.title} ==="]

    max_row = min(ws.max_row or 0, MAX_ROWS_PER_SHEET)
    max_col = min(ws.max_column or 0, MAX_COLS_PER_SHEET)

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
        cleaned = [_clean_cell(cell) for cell in row]
        if any(cleaned):
            lines.append(" | ".join(cleaned))

    return "\n".join(lines).strip()


def read_xlsx(path: Path) -> str:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    parts = []

    for ws in wb.worksheets:
        sheet_text = _sheet_to_text(ws)
        if sheet_text:
            parts.append(sheet_text)

    return "\n\n".join(parts).strip()


def read_csv_file(path: Path) -> str:
    lines = [f"=== CSV: {path.name} ==="]

    with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as f:
        reader = csv.reader(f)
        for idx, row in enumerate(reader):
            if idx >= MAX_ROWS_PER_SHEET:
                break
            cleaned = [_clean_cell(cell) for cell in row[:MAX_COLS_PER_SHEET]]
            if any(cleaned):
                lines.append(" | ".join(cleaned))

    return "\n".join(lines).strip()


def parse_spreadsheet_file(file_path: str | Path) -> ParsedSpreadsheet:
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        text = read_xlsx(path)
        file_type = "xlsx"
    elif suffix == ".csv":
        text = read_csv_file(path)
        file_type = "csv"
    else:
        raise ValueError(f"Unsupported spreadsheet file type: {suffix}")

    created_at = datetime.fromtimestamp(path.stat().st_mtime)

    return ParsedSpreadsheet(
        title=path.name,
        created_at=created_at,
        author=None,
        participants=[],
        raw_path=str(path),
        text=text,
        file_type=file_type,
    )